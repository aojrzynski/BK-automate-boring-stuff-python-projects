#! python3
# cellInverter.py - This program inverts the row and column of the cells in a specified spreadsheet. For example, 
# the value at row 5, column 3 will be at row 3, column 5 (and vice versa).
#
# NOTES: Run this program from the command line, and as an argument provide the filename you want this program to work on.
#         The file should be in the same location as this program.

import openpyxl, sys

# Sets up the workbook object from the excelFile variable. 
# Throws an exception if the filename provided by command line is not found.
try:
    wb = openpyxl.load_workbook(sys.argv[1])
except:
    print('File not found. Please provide a valid filename.')
    sys.exit()
ws = wb.active
# Creates a new workbook object.
newWb = openpyxl.Workbook()
newWs = newWb.active

# For each cell in 'ws', copy the cell coordinates and value into 'newWs', with the coordinates inverted.
for row in ws.iter_rows():
    for cell in row:
        newWs.cell(row=cell.column, column=cell.row, value=cell.value) # Row is the cell.column, and column is th cell.row. Value stays the same.

newWb.save(sys.argv[1])


        