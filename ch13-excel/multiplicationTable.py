#! python3
# multiplicationTable.py - Using a number provided as a command line argument, 
# creates a multiplication table based on that number in an excel file called 'multiplication.xlsx'.
#
# NOTES: Use the command line to access this file, with one argument (any valid number).

import sys, openpyxl
from openpyxl.utils import get_column_letter

# Assigns the argument from the command line to a variable, and makes it an int. 
# If this throws an error, an exception is called and the progrma stops.
try:
    num = int(sys.argv[1]) + 1
except:
    print('That is not a number, please enter a number.')
    sys.exit()

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Multiplication Table'

# Adds the column and row headings.
for i in range(1, num):
    ws.cell(row=1, column=i + 1, value=i)
    ws.cell(row=i + 1, column=1, value=i)

# For every cell, does the math by multiplying the column and row header and writes the answer down in the cell.
for row in ws.iter_rows(min_row=2, min_col=2, max_row=num, max_col=num):
    for cell in row:
        cellX = ws['%s%s' % (get_column_letter(cell.column), '1')]
        cellY = ws['%s%s' % ('A', str(cell.row))]
        cell.value = cellX.value * cellY.value
        
wb.save('multiplication.xlsx')

